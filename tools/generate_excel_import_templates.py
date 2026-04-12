#!/usr/bin/env python3
"""
Writes Supabase import templates as .xlsx (open in Microsoft Excel).
Sheet1 column headers match ALL columns on public.institutes, public.batches,
public.students per supabase/migrations (001, 003, 004, 011).
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "excel_templates"

# --- Full DB column order (matches migrations) ---

INSTITUTES_COLUMNS = [
    "id",
    "institute_code",
    "name",
    "location",
    "address",
    "city",
    "district",
    "taluka",
    "state",
    "country",
    "pincode",
    "mobile_no",
    "is_active",
    "user_count",
    "student_count",
    "last_user_added",
    "sr_no_migration_completed",
    "sr_no_migration_date",
    "sr_no_migration_count",
    "created_at",
    "updated_at",
    "batch_open_time",
    "batch_close_time",
    "batch_duration_minutes",
    "batch_timing_updated_at",
]

BATCHES_COLUMNS = [
    "id",
    "institute_id",
    "name",
    "year",
    "timing",
    "subjects",
    "student_count",
    "created_by",
    "created_at",
    "semester",
    "start_time",
    "end_time",
    "batch_duration_minutes",
    "is_auto_generated",
    "updated_at",
]

STUDENTS_COLUMNS = [
    "id",
    "institute_id",
    "name",
    "first_name",
    "middle_name",
    "last_name",
    "phone_number",
    "sr_no",
    "user_id",
    "year",
    "batch_id",
    "face_embedding",
    "photo_url",
    "created_at",
    "updated_at",
    "sr_no_migrated_at",
    "email",
    "batch_ids",
    "batch_name",
    "batch_timing",
    "subject",
    "subjects",
    "semester",
    "semester_name",
    "role",
    "status",
    "has_device",
    "face_photo_url",
    "uid",
]


def autosize_columns(ws, max_width=55):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max(len(str(c.value or "")) for c in col) + 2, max_width)
        ws.column_dimensions[letter].width = width


def add_instructions(ws, lines: list[str]):
    ws["A1"] = "Instructions (read before filling Sheet1)"
    ws["A1"].font = Font(bold=True, size=12)
    for i, line in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=line)
        ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # --- Institutes (every DB column) ---
    wb_i = Workbook()
    ws = wb_i.active
    ws.title = "institutes"
    ws.append(INSTITUTES_COLUMNS)
    for c in ws[1]:
        c.font = Font(bold=True)

    # Row 2: MSCE001 — leave timestamps/json empty for DB defaults on import
    ws.append(
        [
            "MSCE001",
            "MSCE001",
            "Example Institute Name",
            "Campus A",
            "123 Main Road",
            "Pune",
            "Pune",
            "Haveli",
            "Maharashtra",
            "India",
            "411001",
            "9876543210",
            "TRUE",
            "0",
            "0",
            "",  # last_user_added
            "FALSE",  # sr_no_migration_completed
            "",  # sr_no_migration_date
            "",  # sr_no_migration_count
            "",  # created_at
            "",  # updated_at
            "",  # batch_open_time JSON e.g. {"hour":8,"minute":0}
            "",  # batch_close_time
            "",  # batch_duration_minutes (empty = DB default 60)
            "",  # batch_timing_updated_at
        ]
    )
    ws.append(
        [
            "MSCE002",
            "MSCE002",
            "Second Example College",
            "Block B",
            "456 Station Road",
            "Mumbai",
            "Mumbai",
            "Andheri",
            "Maharashtra",
            "India",
            "400058",
            "9123456780",
            "TRUE",
            "0",
            "0",
            "",
            "FALSE",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    autosize_columns(ws)

    ws_help = wb_i.create_sheet("README", 1)
    add_instructions(
        ws_help,
        [
            "Import into: public.institutes",
            "Sheet1 header row = EVERY column on public.institutes (see supabase/migrations).",
            "Required for new rows: id, institute_code, name (id and institute_code should match).",
            "Leave empty: last_user_added, sr_no_migration_*, created_at, updated_at (DB can default).",
            "batch_open_time / batch_close_time: JSON in Postgres e.g. {\"hour\":8,\"minute\":0} — set in SQL/app if CSV import fails.",
            "After institutes → batches → students.",
        ],
    )
    wb_i.save(OUT_DIR / "institutes_import_template.xlsx")

    # --- Batches (every DB column); id empty = Supabase generates UUID ---
    wb_b = Workbook()
    ws = wb_b.active
    ws.title = "batches"
    ws.append(BATCHES_COLUMNS)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.append(
        [
            "",  # id — leave blank to auto-generate
            "MSCE001",
            "Batch 1 (08:00 - 09:00)",
            "2025",
            "08:00 - 09:00",
            "{}",
            "0",
            "excel_import",
            "",  # created_at
            "1",
            "",  # start_time JSON
            "",  # end_time JSON
            "",  # batch_duration_minutes
            "FALSE",
            "",  # updated_at
        ]
    )
    ws.append(
        [
            "",
            "MSCE001",
            "Batch 2 (09:00 - 10:00)",
            "2025",
            "09:00 - 10:00",
            "{English,Mathematics}",
            "0",
            "excel_import",
            "",
            "1",
            '{"hour":9,"minute":0}',
            '{"hour":10,"minute":0}',
            "60",
            "FALSE",
            "",
        ]
    )
    autosize_columns(ws)
    ws_help = wb_b.create_sheet("README", 1)
    add_instructions(
        ws_help,
        [
            "Import into: public.batches",
            "Sheet1 = ALL columns on public.batches.",
            "id: leave EMPTY for new rows so Postgres generates UUID; then copy id into students.batch_id.",
            "subjects: Postgres text[] — {A,B} or {}.",
            "start_time / end_time: JSON {\"hour\":9,\"minute\":0} if your import tool accepts it; else set in Supabase SQL Editor.",
            "created_at / updated_at: leave empty for default now() where applicable.",
        ],
    )
    wb_b.save(OUT_DIR / "batches_import_template.xlsx")

    # --- Students (every DB column) ---
    wb_s = Workbook()
    ws = wb_s.active
    ws.title = "students"
    ws.append(STUDENTS_COLUMNS)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.append(
        [
            "",  # id auto
            "MSCE001",
            "राजेश पाटील",
            "",
            "",
            "",
            "9123456789",
            "1",
            "ROLL001",
            "2025",
            "PASTE_UUID_BATCH_1_HERE",
            "",  # face_embedding jsonb — app
            "",  # photo_url
            "",  # created_at
            "",  # updated_at
            "",  # sr_no_migrated_at
            "",
            "",  # batch_ids text[] e.g. {uuid} if multi-batch
            "Batch 1 (08:00 - 09:00)",
            "08:00 - 09:00",
            "",  # subject legacy single field
            "{}",
            "1",
            "Semester 1 - 2025",
            "student",
            "approved",
            "FALSE",
            "",  # face_photo_url
            "",  # uid — can match id after insert
        ]
    )
    ws.append(
        [
            "",
            "MSCE001",
            "Sneha Kulkarni",
            "Sneha",
            "",
            "Kulkarni",
            "9988776655",
            "2",
            "ROLL002",
            "2025",
            "PASTE_UUID_BATCH_2_HERE",
            "",
            "",
            "",
            "",
            "",
            "sneha.example@email.com",
            "",
            "Batch 2 (09:00 - 10:00)",
            "09:00 - 10:00",
            "",
            "{English,Mathematics}",
            "1",
            "Semester 1 - 2025",
            "student",
            "approved",
            "FALSE",
            "",
            "",
        ]
    )
    autosize_columns(ws)
    ws_help = wb_s.create_sheet("README", 1)
    add_instructions(
        ws_help,
        [
            "Import into: public.students",
            "Sheet1 = ALL columns on public.students (001_initial_schema + 003_students_extras).",
            "id: leave empty to use DB default gen_random_uuid()::text.",
            "batch_id: paste UUID from public.batches (not the batch name).",
            "subject = legacy one string; subjects = text[] {A,B}. You may leave subject empty if subjects is set.",
            "batch_ids: multi-batch as {uuid1,uuid2} or leave empty.",
            "face_embedding, photo_url, face_photo_url: usually empty on import; fill via app.",
            "uid: optional; app often mirrors student id.",
        ],
    )
    wb_s.save(OUT_DIR / "students_import_template.xlsx")

    print(f"Wrote:\n  {OUT_DIR / 'institutes_import_template.xlsx'}\n  {OUT_DIR / 'batches_import_template.xlsx'}\n  {OUT_DIR / 'students_import_template.xlsx'}")


if __name__ == "__main__":
    main()
